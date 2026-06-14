"""
Unit tests for processors/xlsx_parser.py.

Tests parse_xlsx() using XLSX bytes created in-memory with openpyxl.
All LLM calls are mocked — no real Azure OpenAI calls are made.

Assumptions:
- _llm_serialise_table is imported from pdf_parser into xlsx_parser, so it is
  patched at 'processors.pdf_parser.get_openai_client'.
- XLSX chunks always have page_number == 0 (spreadsheets are not paginated).
- Parent chunk is inserted before its children in the returned list.
- An empty worksheet (no rows) is skipped entirely.
"""
from __future__ import annotations

import io
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from processors.xlsx_parser import parse_xlsx
from shared.models import ChunkType


def _llm_mock():
    mock_client = MagicMock()
    mock_choice = MagicMock()
    mock_choice.message.content = "This table shows employee salary data for the HR department."
    mock_client.chat.completions.create.return_value = MagicMock(choices=[mock_choice])
    return mock_client


def _build_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build XLSX bytes. sheets = {sheet_name: [row_list, ...]}."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_returns_chunks(sample_xlsx_bytes):
    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(sample_xlsx_bytes, "data.xlsx", "https://example.com", "hr", "hr/data.xlsx")
    assert len(chunks) > 0


def test_parse_xlsx_each_sheet_has_parent_chunk():
    file_bytes = _build_xlsx({
        "Employees": [["Name", "Dept"], ["Alice", "HR"], ["Bob", "IT"]],
        "Salaries":  [["Name", "Amount"], ["Alice", 75000]],
    })
    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(file_bytes, "report.xlsx", "https://example.com", "hr", "hr/report.xlsx")
    parents = [c for c in chunks if c.parent_id == ""]
    assert len(parents) == 2


def test_parse_xlsx_parent_chunk_has_sheet_name_as_section_heading():
    file_bytes = _build_xlsx({
        "Q1 Revenue": [["Product", "Revenue"], ["Widget", 50000]],
    })
    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(file_bytes, "report.xlsx", "https://example.com", "hr", "hr/report.xlsx")
    parents = [c for c in chunks if c.parent_id == ""]
    assert len(parents) == 1
    assert parents[0].section_heading == "Q1 Revenue"


def test_parse_xlsx_named_table_creates_table_chunk():
    # Build XLSX with a named Excel table (ListObject)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HR Data"
    ws.append(["Name", "Department"])
    ws.append(["Carol", "Legal"])
    ws.append(["Dan", "IT"])

    from openpyxl.worksheet.table import Table, TableStyleInfo
    tbl = Table(displayName="EmployeeTable", ref="A1:B3")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tbl)

    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(file_bytes, "hr.xlsx", "https://example.com", "hr", "hr/hr.xlsx")

    table_chunks = [c for c in chunks if c.chunk_type == ChunkType.TABLE]
    assert len(table_chunks) >= 1
    assert table_chunks[0].section_subheading == "EmployeeTable"


def test_parse_xlsx_empty_sheet_is_skipped():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    # Leave it truly empty — no rows appended
    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(file_bytes, "empty.xlsx", "https://example.com", "hr", "hr/empty.xlsx")
    assert len(chunks) == 0


def test_parse_xlsx_parent_inserted_before_children(sample_xlsx_bytes):
    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(sample_xlsx_bytes, "data.xlsx", "https://example.com", "hr", "hr/data.xlsx")
    chunk_index = {c.chunk_id: i for i, c in enumerate(chunks)}
    children = [c for c in chunks if c.parent_id != ""]
    for child in children:
        parent_idx = chunk_index.get(child.parent_id)
        assert parent_idx is not None, f"Parent {child.parent_id!r} not in chunks"
        assert parent_idx < chunk_index[child.chunk_id]


def test_parse_xlsx_page_number_is_zero(sample_xlsx_bytes):
    with patch("processors.pdf_parser.get_openai_client", return_value=_llm_mock()):
        chunks = parse_xlsx(sample_xlsx_bytes, "data.xlsx", "https://example.com", "hr", "hr/data.xlsx")
    assert all(c.page_number == 0 for c in chunks)
