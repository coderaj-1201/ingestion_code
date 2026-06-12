"""
XLSX Parser
===========

Excel is fundamentally tabular — every worksheet is a table.
Strategy:
  - Each sheet → one parent chunk (sheet name = section_heading)
  - Named tables (ListObject) → individual table chunks with NL summary
  - If no named tables, chunk by contiguous data regions
  - Sheet title / workbook title from properties → doc_title
  - Each table chunk: table_raw = markdown, content = LLM NL summary
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from uuid import uuid4

import openpyxl
from openpyxl.utils import get_column_letter

from processors.pdf_parser import _llm_serialise_table
from shared.models import ChunkType, RawChunk

logger = logging.getLogger(__name__)


def _sheet_to_markdown(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 200) -> str:
    """Convert worksheet to markdown table, up to max_rows."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # Use first row as header if it looks like headers (all strings)
    header = rows[0]
    col_count = len(header)
    lines = []
    lines.append("| " + " | ".join(str(c or "") for c in header) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")

    for row in rows[1:max_rows]:
        padded = list(row) + [""] * (col_count - len(row))
        lines.append("| " + " | ".join(str(c or "") for c in padded[:col_count]) + " |")

    return "\n".join(lines)


def _table_range_to_markdown(ws, tbl) -> str:
    """Convert a named Excel table range to markdown."""
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )):
        rows.append("| " + " | ".join(str(c or "") for c in row) + " |")
        if row_idx == 0:
            rows.append("| " + " | ".join(["---"] * len(row)) + " |")
    return "\n".join(rows)


def parse_xlsx(
    file_bytes: bytes,
    doc_name: str,
    doc_url: str,
    domain: str,
    blob_path: str,
) -> list[RawChunk]:
    import io
    ingested_at = datetime.now(timezone.utc).isoformat()
    wb          = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    doc_title = doc_name.replace(".xlsx", "").replace("_", " ")
    # Try workbook properties
    if wb.properties and wb.properties.title:
        doc_title = wb.properties.title

    chunks: list[RawChunk] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row == 0:
            continue

        parent_id = str(uuid4())

        # ── Named tables in this sheet ────────────────────────────────────────
        if ws.tables:
            for tbl_name, tbl in ws.tables.items():
                tbl_md = _table_range_to_markdown(ws, tbl)
                if not tbl_md:
                    continue
                nl = _llm_serialise_table(tbl_md, f"{sheet_name} — {tbl_name}")
                chunks.append(RawChunk(
                    chunk_id           = str(uuid4()),
                    parent_id          = parent_id,
                    chunk_type         = ChunkType.TABLE,
                    domain             = domain,
                    doc_name           = doc_name,
                    source             = doc_name,
                    doc_url            = doc_url,
                    file_type          = "xlsx",
                    blob_path          = blob_path,
                    ingested_at        = ingested_at,
                    page_number        = 0,
                    title              = doc_title,
                    section_heading    = sheet_name,
                    section_subheading = tbl_name,
                    content            = nl,
                    table_raw          = tbl_md,
                ))
        else:
            # No named tables — treat whole sheet as one table
            tbl_md = _sheet_to_markdown(ws)
            if not tbl_md:
                continue
            nl = _llm_serialise_table(tbl_md, sheet_name)
            chunks.append(RawChunk(
                chunk_id           = str(uuid4()),
                parent_id          = parent_id,
                chunk_type         = ChunkType.TABLE,
                domain             = domain,
                doc_name           = doc_name,
                source             = doc_name,
                doc_url            = doc_url,
                file_type          = "xlsx",
                blob_path          = blob_path,
                ingested_at        = ingested_at,
                page_number        = 0,
                title              = doc_title,
                section_heading    = sheet_name,
                section_subheading = "",
                content            = nl,
                table_raw          = tbl_md,
            ))

        # Parent chunk — sheet-level summary
        # Build children list for this sheet (already appended above)
        sheet_children = [c for c in chunks if c.parent_id == parent_id]
        sheet_content  = f"Sheet: {sheet_name}. " + (
            sheet_children[0].content if sheet_children else "No data."
        )
        parent_chunk = RawChunk(
            chunk_id           = parent_id,
            parent_id          = "",
            chunk_type         = ChunkType.HEADING,
            domain             = domain,
            doc_name           = doc_name,
            source             = doc_name,
            doc_url            = doc_url,
            file_type          = "xlsx",
            blob_path          = blob_path,
            ingested_at        = ingested_at,
            page_number        = 0,
            title              = doc_title,
            section_heading    = sheet_name,
            section_subheading = "",
            content            = sheet_content,
        )
        # Insert parent before its children so ordering is parent → children
        insert_at = len(chunks) - len(sheet_children)
        chunks.insert(insert_at, parent_chunk)

    logger.info("XLSX parsed: %s → %d chunks", doc_name, len(chunks))
    return chunks
